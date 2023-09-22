import os
import pandas as pd
import requests
from openpyxl import Workbook
import json
from datetime import datetime

print("|==================================|")
print("|== Script iniciado com sucesso! ==|")
print("|==================================|")

import openpyxl
from openpyxl.utils import get_column_letter
nome_arquivo = r"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Date.xlsx"

try:
    workbook = openpyxl.load_workbook(nome_arquivo)
except FileNotFoundError:
    workbook = Workbook()

sheet = workbook.active if "Sheet" in workbook.sheetnames else workbook.create_sheet("Sheet")
sheet["A1"] = "Data"
coluna = get_column_letter(1)
data_atual = datetime.now()
data_formatada = data_atual.strftime("%Y-%m-%d %H:%M:%S")
sheet[f"{coluna}2"] = data_formatada
workbook.save(nome_arquivo)
workbook.close()

print("|==================================|")
print("|== Data adquirida com sucesso! ===|")
print("|==================================|")

hoje1 = datetime.today()
hoje2 = hoje1.strftime('%d-%m %H-%M')
nome = hoje2 + '.xlsx'

with open(r'C:\Users\Iago Piai\Desktop\StatusBolt\Autorization_Infra.txt', 'r') as arquivo:
    autorization = arquivo.read()

headers = {'Authorization': autorization}
html = requests.get(url='https://infra.api.ibbx.tech/monitorator/disconnected-positions?limit=10&page=1', headers=headers, timeout=50)
output = json.loads(html.text)
pages = output['pagination']['totalPages']
total_elements = output['pagination']['totalElements']

wb = Workbook()
ws = wb.active 
ws.append(['Empresa', 'Unidade', 'Position', 'Motivo' , 'Data Entrada', 'Id', 'Chamado', 'Data Abertura', 'Status'])

desabilitados = {

}

hoje1 = datetime.today()
hoje2 = hoje1.strftime('%d/%m')

a = 0
b = 0
while a <= pages:
    a += 1
    headers = {'Authorization': autorization}
    html = requests.get(url=f'https://infra.api.ibbx.tech/monitorator/disconnected-positions?limit=10&page={a}', headers=headers, timeout=50)
    output = json.loads(html.text)
    if 'data' in output:
        for item in output['data']:
            empresa = item['companyName']
            if empresa in desabilitados:
                pass
            else:        
                b += 1
                print(f"Dados do {b}° sensor coletado!")
                empresa = item['companyName']
                unidade = item['facilityName']
                tempo = item['lastAcquisitionDate']
                id = item['activatorId']
                position = item['positionId']
                suporte = item['motive']

                data_str = tempo
                if data_str is not None:
                    data_formatada = datetime.strptime(data_str, "%Y-%m-%dT%H:%M:%S.%fZ")
                    last_date = data_formatada.strftime("%d/%m/%Y")
                else:
                    last_date = 0

                numero_chamado = None
                data_chamado = None
                status = None
                responsavel = None
                

                ws.append([empresa, unidade, position, suporte, hoje2, id, numero_chamado, data_chamado, status, responsavel])

wb.save(fr"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Historico Pontos\{nome}")

print("|==================================|")
print("|=== Arquivo salvo com sucesso! ===|")
print("|==================================|")

df1 = pd.read_excel(r'C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Ativos.xlsx', sheet_name='PontosOff')
df2 = pd.read_excel(fr"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Historico Pontos\{nome}")
df3 = pd.read_excel(r'C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Ativos.xlsx', sheet_name='Adicionados')
df4 = pd.read_excel(r'C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Ativos.xlsx', sheet_name='Removidos')

missing_positions = set(df2['Position']) - set(df1['Position'])
missing_rows = df2[df2['Position'].isin(missing_positions)]

df1 = pd.concat([df1, missing_rows])

positions_to_remove = set(df1['Position']) - set(df2['Position'])
df_removed = df1[df1['Position'].isin(positions_to_remove)]
df1 = df1[~df1['Position'].isin(positions_to_remove)]

df3 = pd.concat([df3, missing_rows])

def concatenate_xlsx_files(folder_path, output_filename):
    if not os.path.exists(folder_path):
        print(f"O caminho '{folder_path}' não existe.")
        return

    xlsx_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]
    
    data_frames = []
    
    for file in xlsx_files:
        file_path = os.path.join(folder_path, file)
        df = pd.read_excel(file_path)
        data_frames.append(df)

    concatenated_df = pd.concat(data_frames, ignore_index=True)

    output_path = os.path.join(folder_path, output_filename)
    concatenated_df.to_excel(output_path, index=False)

if __name__ == "__main__":
    folder_path = r"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Historico Pontos"
    output_filename = r"C:\Users\Iago Piai\Desktop\StatusBolt\Main\Concat.xlsx"
    concatenate_xlsx_files(folder_path, output_filename)

df5 = pd.read_excel(fr"C:\Users\Iago Piai\Desktop\StatusBolt\Main\Concat.xlsx")

print("|==================================|")
print("|===== Arquivos manipulados! ======|")
print("|==================================|")

print("|======================================|")
print("|===== Adquirindo bolts offline! ======|")
print("|======================================|")

import datetime

url = 'https://infra.api.ibbx.tech/monitorator/offline-bolts?limit=10&page=1'
headers = {'Authorization': autorization}
html = requests.get(url=url, headers=headers, timeout=50)
output = json.loads(html.text)
result = []

excluded_bolts = {'F034AA2907E70009000832D687353751','4A04E85707E700090008EE064FE1C617','B212A63A07E700090008E826E2BF5405','AC1F09FFFE0920A8', '43D69CD507E7000500088969B20B8FC4'}

excluded_companies = {}

result = [item for item in output if item['gatewayId'] not in excluded_bolts and item['companyName'] not in excluded_companies]

wb = Workbook()
ws = wb.active

ws.append(['Empresa', 'Unidade', 'Bolt Name', 'Chip', 'Ultima Transmissão', 'Id', 'Status', 'Responsavel'])

try:
    with open("itens_notificados.json", "r") as file:
        itens_notificados = set(json.load(file))
except FileNotFoundError:
    itens_notificados = set()

for item in result:
    name = item['name']
    print(item)
    time_json = item['healthCheckDetails'].get('gatewayRxTime')
    
    if time_json is not None:
        gateway_rx_time_not_formatted = datetime.datetime.fromtimestamp(time_json / 1000)
        gateway_rx_time = gateway_rx_time_not_formatted.strftime("%d/%m/%Y %H:%M")

        current_time = datetime.datetime.now()
        time_difference = current_time - gateway_rx_time_not_formatted
        half_hour = datetime.timedelta(minutes=30)

        if (time_difference >= half_hour or time_difference <= -half_hour) and item['gatewayId'] not in itens_notificados:
            id = item['gatewayId']
            content = item['name']
            facilities_name = item['facilityName']
            companie_name = item['companyName']
            if companie_name in excluded_companies or id in excluded_bolts:
                pass
            else:
                chip = item["healthCheckDetails"]["mobileNetwork"].get("operator")
                if chip is not None:
                    chip = item["healthCheckDetails"]["mobileNetwork"]["operator"]
                    if not chip:
                        chip = 0

                    ws.append([companie_name, facilities_name, content, chip, gateway_rx_time, id])

wb.save(fr"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Historico Bolts\Bolts {nome}")

df6 = pd.read_excel(fr"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Historico Bolts\Bolts {nome}")

pontos_off = df1.shape[0]
bolts_off = df6.shape[0]

with pd.ExcelWriter(r'C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Ativos.xlsx') as writer:
    df6.to_excel(writer, sheet_name='BoltsOff', index=False)
    df1.to_excel(writer, sheet_name='PontosOff', index=False)
    df3.to_excel(writer, sheet_name='Adicionados', index=False)
    df_removed.to_excel(writer, sheet_name='Removidos', index=False)
    df5.to_excel(writer, sheet_name='Historico', index=False)

df_removed.to_excel(fr"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Historico Reconectados\Removed {nome}")

print("|==================================|")
print("|===== Adquirindo sensores R! =====|")
print("|==================================|")

from datetime import datetime

wb = Workbook()
ws = wb.active 

wa = Workbook()
wo = wa.active
wo.title = 'Preventiva'
wo.title = 'Baterias'

with open(r'C:\Users\Iago Piai\Desktop\StatusBolt\Autorization_Infra.txt', 'r') as arquivo:
    autorization = arquivo.read()

headers = {'Authorization': autorization}
html = requests.get(url='https://infra.api.ibbx.tech/summary/endpoints?limit=10&page=1', headers=headers, timeout=50)
output = json.loads(html.text)
pages = output['pagination']['totalPages']
i = 1
a = 1

ws.append(['Empresa', 'Unidade', 'Equipamento', 'Ponto', 'Position', 'Tipo', 'Ativado', 'Conectado', 'Qualidade', 'Bat Inicial', 'Bat Atual', 'Temp', 'Rssi', 'Data', 'Id'])
wo.append(['Empresa', 'Unidade', 'Equipamento', 'Ponto', 'Position', 'Tipo', 'Ativado', 'Conectado', 'Qualidade', 'Bat Inicial', 'Bat Atual', 'Temp', 'Rssi', 'Data', 'Id'])

total_ativados = 0
while a <= pages:
    html = requests.get(url=f'https://infra.api.ibbx.tech/summary/endpoints?limit=10&page={a}', headers=headers, timeout=50)
    output = json.loads(html.text)
    pages = output['pagination']['totalPages']
    a += 1

    for item in output['data']:
        if 'isActivated' in item:
            if item['isActivated'] == True:
                total_ativados += 1

                print(f'Coletandos dados do position: {i}')
                i += 1
                empresa = item['companyName']
                unidade = item['facilityName']
                equipamento = item['assetName']
                ponto = item['positionName']
                position = item['positionId']

                if item['isActivated'] == True:
                    total_ativados += 1

                tipo_sensor = item['sysSensorTypeName']
                if tipo_sensor == 'Vibração e temperatura':
                    tipo_sensor = 'Spectra'
                elif tipo_sensor == 'Analógico Modular':
                    tipo_sensor = 'Modular'
                elif tipo_sensor == 'Energia Elétrica':
                    tipo_sensor = 'Elétrico'
                
                temperatura = item['temperature']
                if temperatura is None:
                    temperatura = 0
                
                if 'initialBatteryVoltage' in item and item['initialBatteryVoltage'] is not None:
                    bateria_inicial = round(item['initialBatteryVoltage'], 1)
                else:
                    bateria_inicial = 0

                if 'batteryVoltage' in item:
                    bateria_atual = item['batteryVoltage']
                    if bateria_atual is not None:
                        o = 1
                    else:
                        bateria_atual = 0
                else:
                    bateria_atual = 0

                qualidade = item['signalQuality']
                if qualidade != '100' and qualidade != '0':
                    qualidade = round(qualidade, 1) 

                rssi = item['lastCollectRSSI']
                if rssi is None:
                    rssi = 0
                
                ativado = item['isActivated']
                if ativado == True:
                    ativado = 'Sim'
                else:
                    ativado = 'Não'

                conectado = item['isConnected']
                if conectado == True:
                    conectado = 'Sim'
                else:
                    conectado = 'Não'
                    
                if 'initialBatteryVoltage' in item:
                    bat_incial = item['initialBatteryVoltage']
                else:
                    bat_incial = 'None'


                temp = item['lastAcquisitionDate']
                if item['lastAcquisitionDate'] is not None:
                    data_hora = datetime.fromtimestamp(temp / 1000)
                    tempo = data_hora.strftime("%d/%m/%Y %H:%M")

                idd = item['boardId']

                if isinstance(bat_incial, (float, int)) and bat_incial > 4.7:
                    ws.append([empresa, unidade, equipamento, ponto, position, tipo_sensor, ativado, conectado, qualidade, round(bateria_atual, 1), bateria_atual, temperatura, rssi, tempo, idd])

                if isinstance(bateria_atual, (float, int)) and 3.2 <= bateria_atual <= 3.7:
                    wo.append([empresa, unidade, equipamento, ponto, position, tipo_sensor, ativado, conectado, qualidade, round(bateria_atual, 1), bateria_atual, temperatura, rssi, tempo, idd])

wa.save(rf"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Bateria_Preventivo.xlsx")
wb.save(rf"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Bateria.xlsx")

print("|==================================|")
print("|===== Finalizada a extração! =====|")
print("|==================================|")

wy = Workbook()
wr = wy.active 
wr.append(['Pcnt Desconectados', 'Pcnt Bolt Off',])
percentagem_pontos_off = (pontos_off / total_ativados) * 1000
percentagem_pontos_off_str = str(round(percentagem_pontos_off, 1)) + '%'

with open(r'C:\Users\Iago Piai\Desktop\StatusBolt\Autorization_Infra.txt', 'r') as arquivo:
    autorization = arquivo.read()

headers = {'Authorization': autorization}
html = requests.get(url='https://infra.api.ibbx.tech/configurator/gateways?limit=10&page=1', headers=headers, timeout=150)
output = json.loads(html.text)
total_gateways = 0
for item in output:
    total_gateways += 1

percentagem_bolts_off = (bolts_off / total_gateways) * 1000
percentagem_bolts_offf = str(round(percentagem_bolts_off, 1)) + '%'
wr.append([percentagem_pontos_off_str, percentagem_bolts_offf])
wr.title = 'Infos'

wy.save(rf"C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Informações.xlsx")

print("|===============================|")
print("|===== Montando Histórico! =====|")
print("|===============================|")

import pandas as pd
import os
import openpyxl

# Diretório onde estão os arquivos
diretorio = r'C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Historico Pontos'

# Lista para armazenar os resultados
resultados = []

# Loop pelos arquivos no diretório
for arquivo in os.listdir(diretorio):
    # Verifica se é um arquivo XLSX e não um diretório
    if os.path.isfile(os.path.join(diretorio, arquivo)) and arquivo.endswith('.xlsx'):
        # Remove a extensão ".xlsx" do nome do arquivo
        nome_arquivo = os.path.splitext(arquivo)[0]
        # Abre o arquivo XLSX
        wb = openpyxl.load_workbook(os.path.join(diretorio, arquivo))
        # Seleciona a primeira planilha (você pode ajustar isso conforme necessário)
        planilha = wb.active
        # Conta as linhas na planilha
        num_linhas = len(list(planilha.iter_rows()))
        # Converta 'nome_arquivo' para o tipo de data e ajuste o ano para 2023
        data = pd.to_datetime(nome_arquivo, format="%d-%m %H-%M").replace(year=2023)
        # Adicione o resultado à lista
        resultados.append({'Data': data, 'Pontos': num_linhas})

# Cria um DataFrame com os resultados
df = pd.DataFrame(resultados)

# Salva o DataFrame em um arquivo Excel
df.to_excel(r'C:\Users\Iago Piai\Desktop\StatusBolt\Recalls\Hist_Pontos.xlsx', index=False)

print("|=================================|")
print("|===== Processos finalizados =====|")
print("|=================================|")